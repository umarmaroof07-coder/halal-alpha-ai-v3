"""
Performance Report — Excel workbook + text summary for backtest results.

Pass/Fail rules:
  1. CAGR > SPY CAGR
  2. Sharpe ≥ 0.5
  3. Sortino ≥ 0.7
  4. Max Drawdown not worse than SPY by more than 10 pp
       → PASS if portfolio_max_dd ≥ spy_max_dd - 0.10
  5. Win Rate ≥ 45%
  6. Out-sample CAGR > 0%
  7. Out-sample beats SPY

Extra warning (not a fail): portfolio max drawdown < -40%.

Both survivorship-bias and look-ahead-bias warnings always appear.
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backtester.engine import (
    BacktestResult,
    SURVIVORSHIP_BIAS_WARNING,
    FUNDAMENTALS_NOT_POINT_IN_TIME_WARNING,
)

REPORTS_DIR = Path("data/reports")

# ---------------------------------------------------------------------------
# Pass/Fail thresholds
# ---------------------------------------------------------------------------

PASS_THRESHOLDS = {
    "sharpe_min":              0.5,
    "sortino_min":             0.7,
    "max_dd_vs_spy_buffer":    0.10,   # portfolio must not be >10pp worse than SPY
    "max_dd_absolute_warn":   -0.40,   # extra warning if worse than -40%
    "win_rate_min":            0.45,
    "out_sample_cagr_min":     0.0,
}


@dataclass
class PassFailResult:
    label: str
    passed: bool
    value: str       # formatted actual value
    threshold: str   # formatted threshold / comparison


def evaluate_pass_fail(result: BacktestResult) -> list[PassFailResult]:
    """
    Evaluate all pass/fail criteria against a BacktestResult.
    Returns list of PassFailResult in display order.
    """
    m   = result.metrics
    bc  = result.benchmark_comparison
    ov  = result.overfitting_report
    thr = PASS_THRESHOLDS

    checks: list[PassFailResult] = []

    # 1. CAGR > SPY
    checks.append(PassFailResult(
        label="CAGR beats SPY",
        passed=m.cagr > bc.spy_cagr,
        value=f"{m.cagr*100:.1f}%",
        threshold=f"SPY {bc.spy_cagr*100:.1f}%",
    ))

    # 2. Sharpe ≥ 0.5
    checks.append(PassFailResult(
        label="Sharpe ≥ 0.5",
        passed=m.sharpe >= thr["sharpe_min"],
        value=f"{m.sharpe:.2f}",
        threshold="≥ 0.50",
    ))

    # 3. Sortino ≥ 0.7
    checks.append(PassFailResult(
        label="Sortino ≥ 0.7",
        passed=m.sortino >= thr["sortino_min"],
        value=f"{m.sortino:.2f}",
        threshold="≥ 0.70",
    ))

    # 4. Max Drawdown vs SPY (main rule)
    dd_pass_threshold = bc.spy_max_drawdown - thr["max_dd_vs_spy_buffer"]
    checks.append(PassFailResult(
        label="Max DD not >10pp worse than SPY",
        passed=m.max_drawdown >= dd_pass_threshold,
        value=f"{m.max_drawdown*100:.1f}%",
        threshold=f"≥ {dd_pass_threshold*100:.1f}%  (SPY {bc.spy_max_drawdown*100:.1f}% − 10pp)",
    ))

    # Extra warning: absolute -40% (not a pass/fail, but flag it)
    if m.max_drawdown < thr["max_dd_absolute_warn"]:
        checks.append(PassFailResult(
            label="⚠ Max DD < −40% (extra warning)",
            passed=False,
            value=f"{m.max_drawdown*100:.1f}%",
            threshold="Warning threshold: −40%",
        ))

    # 5. Win Rate ≥ 45%
    checks.append(PassFailResult(
        label="Win Rate ≥ 45%",
        passed=m.win_rate >= thr["win_rate_min"],
        value=f"{m.win_rate*100:.1f}%",
        threshold="≥ 45%",
    ))

    # 6. Out-sample CAGR > 0%
    checks.append(PassFailResult(
        label="Out-sample CAGR > 0%",
        passed=ov.out_sample.portfolio_cagr > thr["out_sample_cagr_min"],
        value=f"{ov.out_sample.portfolio_cagr*100:.1f}%",
        threshold="> 0%",
    ))

    # 7. Out-sample beats SPY
    checks.append(PassFailResult(
        label="Out-sample beats SPY",
        passed=ov.out_sample.portfolio_cagr > ov.out_sample.spy_cagr,
        value=f"{ov.out_sample.portfolio_cagr*100:.1f}%",
        threshold=f"SPY {ov.out_sample.spy_cagr*100:.1f}%",
    ))

    return checks


def generate_text_summary(result: BacktestResult) -> str:
    """
    Return a formatted multi-line text summary of the backtest.
    Suitable for terminal output and saving as .txt.
    Warnings are always the first section.
    """
    m   = result.metrics
    bc  = result.benchmark_comparison
    ov  = result.overfitting_report
    pf  = evaluate_pass_fail(result)

    n_pass = sum(1 for c in pf if c.passed and not c.label.startswith("⚠"))
    n_fail = sum(1 for c in pf if not c.passed and not c.label.startswith("⚠"))

    lines: list[str] = []
    sep = "═" * 58

    lines += [
        sep,
        "  HALAL ALPHA AI — BACKTEST SUMMARY",
        sep,
        "",
        "⚠  " + SURVIVORSHIP_BIAS_WARNING,
        "",
        "⚠  " + FUNDAMENTALS_NOT_POINT_IN_TIME_WARNING,
        "",
        sep,
        "  FULL PERIOD PERFORMANCE",
        sep,
        f"  {'Metric':<28} {'Portfolio':>10} {'SPY':>10} {'QQQ':>10}",
        "  " + "─" * 54,
        f"  {'CAGR':<28} {m.cagr*100:>9.1f}%  {bc.spy_cagr*100:>8.1f}%  {bc.qqq_cagr*100:>8.1f}%",
        f"  {'Sharpe':<28} {m.sharpe:>10.2f}  {bc.spy_sharpe:>8.2f}  {bc.qqq_sharpe:>8.2f}",
        f"  {'Sortino':<28} {m.sortino:>10.2f}",
        f"  {'Max Drawdown':<28} {m.max_drawdown*100:>9.1f}%  {bc.spy_max_drawdown*100:>8.1f}%  {bc.qqq_max_drawdown*100:>8.1f}%",
        f"  {'Win Rate':<28} {m.win_rate*100:>9.1f}%",
        f"  {'Alpha vs SPY':<28} {bc.alpha_vs_spy*100:>+9.1f}%",
        f"  {'Alpha vs QQQ':<28} {bc.alpha_vs_qqq*100:>+9.1f}%",
        f"  {'Beta vs SPY':<28} {bc.beta_vs_spy:>10.2f}",
        f"  {'Correlation w/ SPY':<28} {bc.correlation_with_spy:>10.2f}",
        f"  {'Total Return':<28} {m.total_return*100:>9.1f}%",
        f"  {'Months':<28} {m.n_months:>10}",
        "",
        sep,
        "  SUB-PERIOD ANALYSIS",
        sep,
        f"  {'Period':<18} {'Port CAGR':>10} {'SPY CAGR':>10} {'Port DD':>9} {'SPY DD':>9}",
        "  " + "─" * 58,
    ]
    for period in [ov.in_sample, ov.out_sample, ov.crisis_2008, ov.crisis_2020, ov.crisis_2022]:
        lines.append(
            f"  {period.period:<18} {period.portfolio_cagr*100:>9.1f}%"
            f"  {period.spy_cagr*100:>8.1f}%"
            f"  {period.portfolio_max_drawdown*100:>7.1f}%"
            f"  {period.spy_max_drawdown*100:>7.1f}%"
        )

    lines += [
        "",
        f"  Performance decay (out vs in): {ov.performance_decay*100:+.1f}%",
        f"  Avg positions: {ov.concentration.avg_n_positions:.1f}  |  "
        f"Avg HHI: {ov.concentration.avg_hhi:.3f}  |  "
        f"Avg top-1 weight: {ov.concentration.avg_top1_weight*100:.1f}%",
        "",
        sep,
        "  PASS / FAIL",
        sep,
    ]
    for c in pf:
        mark = "✓" if c.passed else "✗"
        lines.append(f"  {mark}  {c.label:<40} {c.value:>8}   (threshold: {c.threshold})")

    lines += [
        "",
        f"  Result: {n_pass} passed / {n_fail} failed",
        "",
        sep,
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

_GREEN  = "00C853"
_RED    = "D50000"
_AMBER  = "FF6F00"
_HEADER = "1A237E"
_WHITE  = "FFFFFF"
_WARN   = "FFF9C4"


def _hdr(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=_WHITE)
    cell.fill = PatternFill("solid", fgColor=_HEADER)
    cell.alignment = Alignment(horizontal="center")


def _warn_cell(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=_WARN)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True)


def _pf_cell(ws, row: int, col: int, passed: bool, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=_GREEN if passed else _RED)
    cell.font = Font(bold=True, color=_WHITE)
    cell.alignment = Alignment(horizontal="center")


def save_excel_report(
    result: BacktestResult,
    output_path: Path | None = None,
) -> Path:
    """
    Write a multi-sheet Excel workbook summarising the backtest.

    Returns the path to the saved file.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        today = date.today().strftime("%Y%m%d")
        output_path = REPORTS_DIR / f"backtest_report_{today}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    m   = result.metrics
    bc  = result.benchmark_comparison
    ov  = result.overfitting_report
    pf  = evaluate_pass_fail(result)

    # ── Sheet 1: Warnings ───────────────────────────────────────────────────
    ws = wb.create_sheet("Warnings")
    _warn_cell(ws, 1, 1, "⚠ " + SURVIVORSHIP_BIAS_WARNING)
    _warn_cell(ws, 3, 1, "⚠ " + FUNDAMENTALS_NOT_POINT_IN_TIME_WARNING)
    ws.column_dimensions["A"].width = 100
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[3].height = 60

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    headers = ["Metric", "Portfolio", "SPY", "QQQ"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, 1, i, h)

    rows = [
        ("CAGR",          f"{m.cagr*100:.1f}%",          f"{bc.spy_cagr*100:.1f}%",         f"{bc.qqq_cagr*100:.1f}%"),
        ("Sharpe",        f"{m.sharpe:.2f}",              f"{bc.spy_sharpe:.2f}",             f"{bc.qqq_sharpe:.2f}"),
        ("Sortino",       f"{m.sortino:.2f}",             "—",                                "—"),
        ("Max Drawdown",  f"{m.max_drawdown*100:.1f}%",   f"{bc.spy_max_drawdown*100:.1f}%",  f"{bc.qqq_max_drawdown*100:.1f}%"),
        ("Win Rate",      f"{m.win_rate*100:.1f}%",       "—",                                "—"),
        ("Alpha vs SPY",  f"{bc.alpha_vs_spy*100:+.1f}%", "—",                                "—"),
        ("Alpha vs QQQ",  f"{bc.alpha_vs_qqq*100:+.1f}%","—",                                "—"),
        ("Beta vs SPY",   f"{bc.beta_vs_spy:.2f}",        "—",                                "—"),
        ("Corr w/ SPY",   f"{bc.correlation_with_spy:.2f}","—",                               "—"),
        ("Total Return",  f"{m.total_return*100:.1f}%",   "—",                                "—"),
        ("Months",        str(m.n_months),                "—",                                "—"),
        ("Best Month",    f"{m.best_month*100:.1f}%",     "—",                                "—"),
        ("Worst Month",   f"{m.worst_month*100:.1f}%",    "—",                                "—"),
        ("Avg Turnover",  f"{m.avg_turnover*100:.1f}%",   "—",                                "—"),
    ]
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # ── Sheet 3: Pass/Fail ──────────────────────────────────────────────────
    ws = wb.create_sheet("Pass_Fail")
    for i, h in enumerate(["Criterion", "Result", "Actual", "Threshold"], 1):
        _hdr(ws, 1, i, h)
    for r, c in enumerate(pf, 2):
        ws.cell(row=r, column=1, value=c.label)
        _pf_cell(ws, r, 2, c.passed, "PASS" if c.passed else "FAIL")
        ws.cell(row=r, column=3, value=c.value)
        ws.cell(row=r, column=4, value=c.threshold)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40

    # ── Sheet 4: Calendar Returns ───────────────────────────────────────────
    ws = wb.create_sheet("Calendar_Returns")
    for i, h in enumerate(["Year", "Portfolio", "SPY", "QQQ"], 1):
        _hdr(ws, 1, i, h)
    port_cy = m.calendar_year_returns
    spy_cy = bc.spy_cagr   # placeholder — real impl would need SPY monthly data split
    for r, (yr, pr) in enumerate(sorted(port_cy.items()), 2):
        ws.cell(row=r, column=1, value=yr)
        ws.cell(row=r, column=2, value=f"{pr*100:.1f}%")
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # ── Sheet 5: Equity Curve ───────────────────────────────────────────────
    ws = wb.create_sheet("Equity_Curve")
    for i, h in enumerate(["Date", "Portfolio", "SPY", "QQQ"], 1):
        _hdr(ws, 1, i, h)
    for r, row in enumerate(result.equity_curve, 2):
        ws.cell(row=r, column=1, value=row["date"])
        ws.cell(row=r, column=2, value=row["value"])
        ws.cell(row=r, column=3, value=row.get("spy_value"))
        ws.cell(row=r, column=4, value=row.get("qqq_value"))
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # ── Sheet 6: Overfitting ────────────────────────────────────────────────
    ws = wb.create_sheet("Overfitting")
    for i, h in enumerate(["Period", "Port CAGR", "SPY CAGR", "Port DD", "SPY DD", "Win Rate", "Months"], 1):
        _hdr(ws, 1, i, h)
    periods = [ov.in_sample, ov.out_sample, ov.crisis_2008, ov.crisis_2020, ov.crisis_2022]
    for r, p in enumerate(periods, 2):
        ws.cell(row=r, column=1, value=p.period)
        ws.cell(row=r, column=2, value=f"{p.portfolio_cagr*100:.1f}%")
        ws.cell(row=r, column=3, value=f"{p.spy_cagr*100:.1f}%")
        ws.cell(row=r, column=4, value=f"{p.portfolio_max_drawdown*100:.1f}%")
        ws.cell(row=r, column=5, value=f"{p.spy_max_drawdown*100:.1f}%")
        ws.cell(row=r, column=6, value=f"{p.win_rate*100:.1f}%")
        ws.cell(row=r, column=7, value=p.n_months)
    ws.cell(row=len(periods)+3, column=1, value="Performance decay (out vs in):")
    ws.cell(row=len(periods)+3, column=2, value=f"{ov.performance_decay*100:+.1f}%")
    # Concentration
    ws.cell(row=len(periods)+5, column=1, value="Avg positions")
    ws.cell(row=len(periods)+5, column=2, value=f"{ov.concentration.avg_n_positions:.1f}")
    ws.cell(row=len(periods)+6, column=1, value="Avg HHI")
    ws.cell(row=len(periods)+6, column=2, value=f"{ov.concentration.avg_hhi:.4f}")
    ws.cell(row=len(periods)+7, column=1, value="Avg top-1 weight")
    ws.cell(row=len(periods)+7, column=2, value=f"{ov.concentration.avg_top1_weight*100:.1f}%")
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # ── Sheet 7: Positions History ──────────────────────────────────────────
    ws = wb.create_sheet("Positions_History")
    for i, h in enumerate(["Date", "Tickers", "Weights", "Portfolio Return", "Turnover"], 1):
        _hdr(ws, 1, i, h)
    for r, snap in enumerate(result.monthly_snapshots, 2):
        ws.cell(row=r, column=1, value=snap.date)
        ws.cell(row=r, column=2, value=", ".join(snap.tickers))
        ws.cell(row=r, column=3, value=", ".join(f"{w*100:.0f}%" for w in snap.weights))
        ws.cell(row=r, column=4, value=f"{snap.portfolio_return*100:.2f}%")
        ws.cell(row=r, column=5, value=f"{snap.turnover*100:.1f}%")
    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 25

    wb.save(output_path)
    return output_path


def save_text_summary(result: BacktestResult, output_path: Path | None = None) -> Path:
    """Save the text summary to a .txt file and return the path."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        today = date.today().strftime("%Y%m%d")
        output_path = REPORTS_DIR / f"backtest_summary_{today}.txt"
    output_path.write_text(generate_text_summary(result), encoding="utf-8")
    return output_path
